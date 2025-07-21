from openpyxl import load_workbook
import json

wb = load_workbook("reconstructed.xlsx")
ws = wb.active

data = []
for row in ws.iter_rows(values_only=False):
    row_index = row[0].row  # tất cả cell cùng hàng nên dùng row[0].row
    row_data = []
    for cell in row:
        row_data.append({
            "text": cell.value,
            "link": cell.hyperlink.target if cell.hyperlink else None,
            "raw": repr(cell.value)
        })
    data.append({
        "row": row_index,
        "data": row_data
    })

with open("output_structured.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
