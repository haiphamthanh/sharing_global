import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Load JSON từ file
with open("input.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# Tạo workbook mới
wb = Workbook()
ws = wb.active

# Duyệt từng hàng
for row_idx, row in enumerate(data, start=1):
    for col_idx, cell in enumerate(row, start=1):
        col_letter = get_column_letter(col_idx)
        excel_cell = ws[f"{col_letter}{row_idx}"]

        text = cell.get("text", "")
        link = cell.get("link", None)

        excel_cell.value = text
        excel_cell.alignment = Alignment(
            wrap_text=True)  # Cho phép hiển thị \n

        if link:
            excel_cell.hyperlink = link
            excel_cell.style = "Hyperlink"

# Lưu file Excel
wb.save("reconstructed.xlsx")
